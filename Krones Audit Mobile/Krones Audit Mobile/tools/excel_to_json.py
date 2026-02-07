# tools/excel_to_json.py
from openpyxl import load_workbook
import json
from pathlib import Path

EXCEL = Path("Template-questoes.xlsx")
OUT = Path("app/data/questions.json")
LANG_SHEETS = ["PT-BR", "ES"]  # adicione "EN" no futuro

def read_sheet(ws):
    # Espera colunas: Item | Description | Weighthting | Comments
    rows = []
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            header = [str(c).strip() if c is not None else "" for c in row]
            continue
        data = dict(zip(header, row))
        if not data.get("Item"):
            continue
        rows.append({
            "Item": str(data["Item"]).strip(),
            "Description": (data.get("Description") or "").strip(),
            "Weighthting": float(data.get("Weighthting") or 0)
        })
    return rows

def build_lang_map(rows):
    return {r["Item"]: r["Description"] for r in rows}

def level_of(item_id: str) -> int:
    return item_id.count(".") + 1

def main():
    wb = load_workbook(EXCEL, data_only=True)
    # Carrega todas as línguas como mapas
    lang_maps = {}
    for sheet in LANG_SHEETS:
        ws = wb[sheet]
        lang_maps[sheet] = build_lang_map(read_sheet(ws))

    rows_pt = [{"Item": k, "Description": v, "Weighthting": None}
               for k, v in lang_maps[LANG_SHEETS[0]].items()]
    # Melhor: reler PT-BR com pesos reais
    ws_pt = wb[LANG_SHEETS[0]]
    rows_pt = read_sheet(ws_pt)

    # Monta árvore a partir de PT-BR
    topics = []
    current_topic = None
    current_group = None

    for r in rows_pt:
        item, desc, w = r["Item"], r["Description"], r["Weighthting"]
        lvl = level_of(item)

        # Monta dict multilíngue para o título
        title = {}
        for lang, m in lang_maps.items():
            if item in m:
                title[lang] = m[item]
        # Futuro EN: title["EN"] = ...

        if lvl == 1:
            current_topic = {
                "id": item,
                "weight": w,
                "title": title,
                "groups": []
            }
            topics.append(current_topic)
            current_group = None
        elif lvl == 2:
            current_group = {
                "id": item,
                "title": title,
                "weight": 1,   # no seu Excel é sempre 1
                "questions": []
            }
            current_topic["groups"].append(current_group)
        elif lvl == 3:
            q = {
                "id": item,
                "title": title,
                "weight": w
            }
            if current_group is None:
                # Caso raro: se houver perguntas sem grupo explícito
                current_group = {
                    "id": item.rsplit(".", 1)[0],
                    "title": title,
                    "weight": 1,
                    "questions": []
                }
                current_topic["groups"].append(current_group)
            current_group["questions"].append(q)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    data = {"languages": LANG_SHEETS + (["EN"] if "EN" in wb.sheetnames else []),
            "topics": topics}
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Gerado: {OUT.resolve()}")

if __name__ == "__main__":
    main()